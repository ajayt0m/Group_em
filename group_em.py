
import os
import random
import openpyxl

# read list of students from user input
students = input("Enter the names of students separated by commas: ").split(',')

# get number of members per group required from user
members_per_group = int(input("Enter the number of members per group required: "))

while True:
    # shuffle the list of students randomly
    random.shuffle(students)

    # split the shuffled list into groups with the specified number of members per group
    groups = [students[i:i+members_per_group] for i in range(0, len(students), members_per_group)]

    # print the groups
    print("The groups are:")
    for i, group in enumerate(groups):
        print(f"Group {i+1}: {', '.join(group)}")

    # ask user if they want to shuffle and regroup again
    answer = input("Do you want to shuffle and regroup again? (y/n): ")
    if answer.lower() == 'n':
        break

# delete file if it exists
if os.path.exists('D:\\GitHub\\Group_em\\Group_emd.xlsx'):
    os.remove('D:\\GitHub\\Group_em\\Group_emd.xlsx')

# export the groups to an Excel file
filename = 'D:\\GitHub\\Group_em\\Group_emd.xlsx'
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = 'Groups'

# write column headers
for i, group in enumerate(groups):
    worksheet.cell(row=1, column=i+1, value=f"Group {i+1}")

# write group members
for i, group in enumerate(groups):
    for j, member in enumerate(group):
        worksheet.cell(row=j+2, column=i+1, value=member)

workbook.save(filename)
print(f"Groups have been exported to {filename}.")


    


